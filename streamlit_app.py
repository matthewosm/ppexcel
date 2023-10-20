import sqlite3
from openpyxl import Workbook
import streamlit as st
import os
import tempfile

def convert_to_excel(db_path, output_name):
    # Connect to the SQLite database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Fetch all table names in the database
    table_list_query = "SELECT name FROM sqlite_master WHERE type='table';"
    cursor.execute(table_list_query)
    tables = cursor.fetchall()

    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # Iterate over all tables
    for table in tables:
        table_name = table[0]
        
        # Create a new sheet for the table
        ws = wb.create_sheet(title=table_name)
        
        # Get the columns (headers)
        cursor.execute(f'SELECT * FROM {table_name} LIMIT 0')
        headers = [description[0] for description in cursor.description]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Write the data from the table to the Excel sheet
        for row_num, row_data in enumerate(cursor.execute(f'SELECT * FROM {table_name}'), 2):
            for col_num, data in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=data)

    # Save the workbook with the specified output name
    wb.save(output_name)
    return output_name

def main():
    st.set_page_config(page_title="Shift Asta Powerproject to Excel Converter", page_icon="https://www.shift-construction.com/wp-content/uploads/2023/10/shift_square_icon.png", layout="centered", initial_sidebar_state="auto", menu_items=None)
    # Add the JPEG header image
    image_path = "https://www.shift-construction.com/wp-content/uploads/2023/10/shift-grey-logo-white-text_small.png"
    st.image(image_path, width=100)
    st.title("Asta Powerproject to Excel Converter")

    # Add link to your website
    website_url = "https://www.shift-construction.com"
    st.markdown(f"Visit [Shift - Digital Construction]({website_url})")

    # Upload the SQLite file
    uploaded_file = st.file_uploader("Choose an Asta Powerproject file", type="pp")
    if uploaded_file:
        # Extract the filename (without extension) from the uploaded file
        uploaded_filename = uploaded_file.name.split('.')[0]
        output_name = os.path.join(tempfile.gettempdir(), f"{uploaded_filename}.xlsx")
        
        # Save the uploaded file temporarily
        temp_db_path = os.path.join(tempfile.gettempdir(), "temp.db")
        with open(temp_db_path, "wb") as f:
            f.write(uploaded_file.getvalue())

        # Add spinner for the conversion process
        with st.spinner("Processing..."):
            excel_path = convert_to_excel(temp_db_path, output_name)

        # Provide a link to download the Excel file
        with open(excel_path, "rb") as f:
            st.download_button(
                label="Download Excel File",
                data=f,
                file_name=f"{uploaded_filename}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()
